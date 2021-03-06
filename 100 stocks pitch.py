import openpyxl

filename = "C:\\Users\shang\OneDrive\Desktop\\100 US Mid-Cap Stocks.xlsx"

workbook = openpyxl.load_workbook(filename)

worksheet1 = workbook.active
insertrows = []
rows = worksheet1.max_row

print(rows)
print(worksheet1.max_column)

for i in range(rows-1,1,-1):
    if worksheet1.cell(row=i+1,column=5).value != worksheet1.cell(row=i,column=5).value:
        insertrows.append(i+1)

insertrows.append(2)

a = 0
ret = 0
vol = 0
vol_sum = 0
count = 0
count2 = 0

daily = 0
daily_sum = 0
count1 = 0
daily_mean = 0
list_return = []
list_vol = []
summary = 0

for i in range(0,len(insertrows)-1,1):

    worksheet1.insert_rows(insertrows[i])

    for t in range(insertrows[i]-1,insertrows[i+1],-1):
        daily = (worksheet1.cell(row=t,column=4).value-worksheet1.cell(row=t-1,column=4).value)/worksheet1.cell(row=t-1,column=4).value
        worksheet1.cell(row=t,column=6).value = daily
        daily_sum += daily
        count1 += 1

    daily_mean = daily_sum/count1
    list_return.append(daily_mean)
    daily_sum = 0
    count1 = 0


    for j in range(insertrows[i]-1,insertrows[i+1],-1):
        summary += worksheet1.cell(row=j,column=6).value
        count += 1
    mean = summary/count
    summary = 0
    count = 0

    for k in range(insertrows[i]-1,insertrows[i+1],-1):
        vol_sum += (worksheet1.cell(row=k,column=6).value-mean)**2
        count2 += 1
    vol = vol_sum/(count2-1)
    vol = vol**(1/2)
    list_vol.append(vol)
    vol_sum = 0
    count2 = 0

    worksheet1.cell(row=insertrows[i],column=1).value = 'return is ' + str(daily_mean)
    worksheet1.cell(row=insertrows[i],column=2).value = 'volatility is ' + str(vol)




workbook.create_sheet('return',1)
worksheet2 = workbook.worksheets[1]
worksheet2.cell(row=1,column=1).value = 'stock index'
worksheet2.cell(row=1,column=2).value = 'return'
worksheet2.cell(row=1,column=3).value = 'volatility'

for i in range(0,100,1):
    worksheet2.cell(row=i+2, column=1).value = 100-i
    worksheet2.cell(row=i+2, column=2).value = list_return[i]
    worksheet2.cell(row=i+2, column=3).value = list_vol[i]



workbook.save(filename=filename)
